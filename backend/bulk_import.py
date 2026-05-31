"""CSV / XLSX bulk-import helpers: file reading + fuzzy header matching.

Pure parsing/normalization only — no DB access. The HTTP layer (routes.py)
layers on organization-scoped validation (duplicates, class-teacher scope) and
the actual record creation.

  * ``read_table``      -> (headers, rows) from .csv or .xlsx bytes
  * ``match_columns``   -> {canonical_field: actual_header} via synonym + fuzzy
  * ``extract_records`` -> per-row {canonical_field: value} dicts

Fuzzy matching handles the examples in the spec:
  "father name" -> parent_name, "mob" -> phone, "adm no" -> admission_number.
"""

import csv
import io
import re
import difflib


# Canonical field -> accepted header synonyms (lower-case, loosely written).
STUDENT_FIELDS = {
    "name": ["name", "full name", "student name", "student", "child name"],
    "parent_name": [
        "parent name", "father name", "fathers name", "father", "guardian",
        "guardian name", "parent", "parent guardian",
    ],
    "mother_name": ["mother name", "mothers name", "mother", "mom name"],
    "email": [
        "email", "e mail", "email id", "mail", "email address",
        "parent email", "parents email", "guardian email",
    ],
    "phone": [
        "phone", "mobile", "mob", "mob no", "mobile no", "mobile number",
        "contact", "contact no", "contact number", "phone number", "ph", "cell",
    ],
    "class": ["class", "grade", "standard", "std", "class grade"],
    "section": ["section", "sec", "div", "division"],
    "roll_no": ["roll no", "roll number", "roll", "rollno", "sr no", "serial no"],
    "admission_number": [
        "admission number", "adm no", "adm", "admission no", "admno",
        "admission", "admission id", "enrollment no", "enrolment no",
    ],
    "date_of_birth": ["date of birth", "dob", "birth date", "d o b", "birthday"],
    "gender": ["gender", "sex", "m f"],
    "address": ["address", "residence", "home address", "city", "addr"],
    "blood_group": ["blood group", "blood", "bg", "blood grp"],
    "admission_date": ["admission date", "joining date", "date of admission", "doj", "enrolled on"],
}

TEACHER_FIELDS = {
    "name": ["name", "full name", "teacher name", "teacher", "employee name"],
    "gender": ["gender", "sex", "m f"],
    "email": ["email", "e mail", "email id", "mail", "email address"],
    "phone": [
        "phone", "mobile", "mob", "mob no", "mobile no", "mobile number",
        "contact", "contact no", "contact number", "phone number", "ph", "cell",
    ],
    "qualification": ["qualification", "qual", "degree", "education", "qualifications"],
    "designation": ["designation", "role", "post", "title", "position"],
    "primary_subject": [
        "primary subject", "main subject", "subject", "teaching subject",
        "primary", "subject 1", "subject1",
    ],
    "secondary_subject": [
        "secondary subject", "other subject", "subject 2", "subject2", "secondary",
    ],
    "joining_date": ["joining date", "date of joining", "doj", "joined on", "joining"],
    "experience_years": [
        "experience", "experience years", "years of experience", "exp",
        "exp years", "total experience", "yrs exp",
    ],
    "availability": ["availability", "employment type", "type", "full part time"],
    "status": ["status", "active", "employment status", "state"],
}

FIELD_SETS = {"students": STUDENT_FIELDS, "teachers": TEACHER_FIELDS}

# Minimum similarity for a fuzzy (non-synonym) header match.
FUZZY_THRESHOLD = 0.82


def _normalize(text):
    """Lower-case, strip punctuation, collapse whitespace for header matching."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def read_table(filename, content):
    """Read a .csv/.xlsx file into (headers, rows).

    ``rows`` is a list of dicts keyed by the original header text. Raises
    ValueError for unsupported types or empty/headerless files.
    """
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        return _read_csv(content)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _read_xlsx(content)
    raise ValueError("Unsupported file type. Please upload a .csv or .xlsx file.")


def _read_csv(content):
    if isinstance(content, bytes):
        # utf-8-sig transparently drops a BOM if Excel added one.
        text = content.decode("utf-8-sig", errors="replace")
    else:
        text = content
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("The file is empty.")
    headers = [(_h or "").strip() for _h in rows[0]]
    return _rows_from_matrix(headers, rows[1:])


def _read_xlsx(content):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    matrix = []
    for row in ws.iter_rows(values_only=True):
        matrix.append(["" if c is None else c for c in row])
    if not matrix:
        raise ValueError("The spreadsheet is empty.")
    headers = [str(h).strip() for h in matrix[0]]
    return _rows_from_matrix(headers, matrix[1:])


def _rows_from_matrix(headers, data_rows):
    if not any(h for h in headers):
        raise ValueError("Could not find a header row in the file.")
    out = []
    for raw in data_rows:
        # Skip fully-blank rows.
        if not any(str(c).strip() for c in raw):
            continue
        record = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            value = raw[i] if i < len(raw) else ""
            record[h] = "" if value is None else str(value).strip()
        out.append(record)
    return headers, out


def match_columns(headers, fields):
    """Map canonical fields to the best-matching header.

    Order of preference: exact synonym match, then fuzzy similarity above
    ``FUZZY_THRESHOLD``. Each header is consumed by at most one field.
    """
    norm_headers = [(h, _normalize(h)) for h in headers if h]
    mapping = {}
    used = set()

    for canon, synonyms in fields.items():
        norm_syn = {_normalize(canon)} | {_normalize(s) for s in synonyms}

        # 1) exact synonym / canonical match
        chosen = None
        for h, nh in norm_headers:
            if h in used:
                continue
            if nh in norm_syn:
                chosen = h
                break

        # 2) fuzzy match against any synonym
        if not chosen:
            best_ratio, best_h = 0.0, None
            for h, nh in norm_headers:
                if h in used or not nh:
                    continue
                ratio = max(difflib.SequenceMatcher(None, nh, ns).ratio() for ns in norm_syn)
                # Also reward token containment ("adm no" within "adm no class").
                if any(ns and (ns in nh or nh in ns) for ns in norm_syn):
                    ratio = max(ratio, 0.9)
                if ratio > best_ratio:
                    best_ratio, best_h = ratio, h
            if best_h and best_ratio >= FUZZY_THRESHOLD:
                chosen = best_h

        if chosen:
            mapping[canon] = chosen
            used.add(chosen)

    return mapping


def extract_records(rows, mapping):
    """Turn raw header-keyed rows into canonical-field dicts.

    Returns a list of ``{"row": <1-based source row>, "data": {canon: value}}``.
    Row numbers start at 2 (row 1 is the header) to match what a user sees in
    their spreadsheet.
    """
    out = []
    for idx, raw in enumerate(rows, start=2):
        data = {}
        for canon, header in mapping.items():
            data[canon] = (raw.get(header) or "").strip()
        out.append({"row": idx, "data": data})
    return out


def unmapped_headers(headers, mapping):
    """Headers present in the file that weren't matched to any canonical field."""
    used = set(mapping.values())
    return [h for h in headers if h and h not in used]
